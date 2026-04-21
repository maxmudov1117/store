from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import ExpressionWrapper, F, FloatField, Q
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.http import HttpResponse, FileResponse
from django.views import View
from .models import ImportProduct, Sale, PayDebt
from main.models import Product, Client, Branch
from .utils import generate_thermal_receipt

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class SaleView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request):
        sales = Sale.objects.filter(branch=request.user.branch).order_by('-created_at', '-id')
        products = Product.objects.filter(branch=request.user.branch).order_by('name')
        clients = Client.objects.filter(branch=request.user.branch).order_by('name')

        # Search
        search = request.GET.get('search', '').strip()
        if search:
            sales = sales.filter(
                Q(product__name__icontains=search) | Q(client__name__icontains=search) | Q(client__shop_name__icontains=search)
            )

        # Date filter
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        if date_from:
            sales = sales.filter(created_at__gte=date_from)
        if date_to:
            sales = sales.filter(created_at__lte=date_to)

        # Pagination
        paginator = Paginator(sales, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Check for automatic receipt popup
        auto_download_id = request.session.pop('last_sale_id', None)

        context = {
            'sales': page_obj,
            'page_obj': page_obj,
            'products': products,
            'clients': clients,
            'search': search,
            'date_from': date_from,
            'date_to': date_to,
            'auto_download_id': auto_download_id,
        }
        return render(request, 'sales.html', context=context)

    def post(self, request):
        if not request.user.branch:
            return render(request, 'warning.html', {
                'warning_title': 'Filial biriktirilmagan',
                'warning_message': "Sizning hisobingizga hech qanday filial (branch) biriktirilmagan. Iltimos, admin panel orqali filial biriktiring.",
                'back_url': 'sales'
            })

        product = get_object_or_404(Product, id=request.POST.get('product_id'))
        client = get_object_or_404(Client, id=request.POST.get('client_id'))
        
        # Xavfsiz float o'girish
        def get_float(val):
            try: return float(val) if val else 0.0
            except: return 0.0

        quantity = get_float(request.POST.get('quantity'))
        total_price = get_float(request.POST.get('total_price'))
        paid_price = get_float(request.POST.get('paid_price'))
        debt_price = get_float(request.POST.get('debt_price'))

        # check product quantity
        context = self.check_enough_product(product, quantity)
        if context is not None:
            return render(request, 'warning.html', context=context)

        # Avtomatik hisoblash mantiqi
        if not total_price or total_price == 0:
            total_price = product.price * quantity

        if paid_price == 0 and debt_price == 0:
            paid_price = total_price
            debt_price = 0.0
        elif paid_price > 0 and debt_price == 0:
            debt_price = total_price - paid_price
        elif paid_price == 0 and debt_price > 0:
            paid_price = total_price - debt_price

        sale = Sale.objects.create(
            product=product,
            client=client,
            quantity=quantity,
            total_price=total_price,
            paid_price=paid_price,
            debt_price=debt_price,
            user=request.user,
            branch=request.user.branch
        )
        # sub product quantity
        product.quantity -= quantity
        product.save()

        # add client debt
        client.debt += debt_price
        client.save()

        # Store last sale id for automatic receipt popup
        request.session['last_sale_id'] = sale.id

        return redirect('sales')

    def check_enough_product(self, product, quantity):
        if product.quantity < quantity:
            warning_message = f"{product.name} so'ralgan miqdor mavjud emas!"
            warning_title = "Maxsulot yetarli emas!"
            back_url = 'sales'
            context = {
                'warning_message': warning_message,
                'warning_title': warning_title,
                'back_url': back_url,
            }
            return context
        return None


class EditSalaView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request, pk):
        sale = get_object_or_404(Sale, branch=request.user.branch, pk=pk)
        products = Product.objects.filter(branch=request.user.branch)
        clients = Client.objects.filter(branch=request.user.branch)

        context = {
            'sale': sale,
            'products': products,
            'clients': clients,
        }
        return render(request, 'edit-sales.html', context=context)

    def post(self, request, pk):
        sale = get_object_or_404(Sale, branch=request.user.branch, pk=pk)

        product = get_object_or_404(Product, id=request.POST.get('product_id'))
        client = get_object_or_404(Client, id=request.POST.get('client_id'))

        def get_float(val):
            try: return float(val) if val else 0.0
            except: return 0.0

        quantity = get_float(request.POST.get('quantity'))
        total_price = get_float(request.POST.get('total_price'))
        paid_price = get_float(request.POST.get('paid_price'))
        debt_price = get_float(request.POST.get('debt_price'))

        if quantity == 0:
            context = {
                'warning_message': 'Miqdor kiritilishi shart!',
                'warning_title': 'Xatolik',
                'back_url': 'sales'
            }
            return render(request, 'warning.html', context=context)

        # Avtomatik hisoblash
        if not total_price or total_price == 0:
            total_price = product.price * quantity

        if paid_price == 0 and debt_price == 0:
            paid_price = total_price
            debt_price = 0.0
        elif paid_price > 0 and debt_price == 0:
            debt_price = total_price - paid_price
        elif paid_price == 0 and debt_price > 0:
            paid_price = total_price - debt_price

        # AVVALGI MA'LUMOTLARNI SAQLASH
        old_product = sale.product
        old_client = sale.client
        old_quantity = sale.quantity
        old_debt_price = sale.debt_price

        # HECH NARSA O'ZGARMAGAN BO'LSA, BAZAGA TEGMASLIK
        if (product.id == old_product.id and
                client.id == old_client.id and
                quantity == old_quantity and
                total_price == sale.total_price and
                paid_price == sale.paid_price and
                debt_price == old_debt_price):
            return redirect('sales')

        # Miqdorni tekshirish
        if product.id == old_product.id:
            quantity_difference = quantity - old_quantity
            if quantity_difference > 0:
                available = old_product.quantity
                if available < quantity_difference:
                    return render(request, 'warning.html', {
                        'warning_message': f"{product.name} yetarli emas!",
                        'warning_title': "Xato",
                        'back_url': 'sales'
                    })
        else:
            if product.quantity < quantity:
                return render(request, 'warning.html', {
                    'warning_message': f"{product.name} yetarli emas!",
                    'warning_title': "Xato",
                    'back_url': 'sales'
                })

        # AVVALGI MAHSULOT MIQDORINI QAYTARISH
        old_product.quantity += old_quantity
        old_product.save()

        # AVVALGI MIJOZ QARZINI KAMAYTIRISH
        if hasattr(old_client, 'debt'):
            old_client.debt -= old_debt_price
            old_client.save()

        # SALE NI YANGILASH
        sale.product = product
        sale.client = client
        sale.quantity = quantity
        sale.total_price = total_price
        sale.paid_price = paid_price
        sale.debt_price = debt_price
        sale.save()

        # YANGI MAHSULOT MIQDORINI KAMAYTIRISH
        product.quantity -= quantity
        product.save()

        # YANGI MIJOZ QARZINI QO'SHISH
        if hasattr(client, 'debt'):
            client.debt += debt_price
            client.save()

        return redirect('sales')


class DeleteSaleView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request, pk):
        sale = get_object_or_404(Sale, pk=pk, branch=request.user.branch)
        context = {
            'sale': sale,
        }
        return render(request, 'delete-sale.html', context=context)

    def post(self, request, pk):
        sale = get_object_or_404(Sale, pk=pk, branch=request.user.branch)

        # Mahsulot miqdorini qaytarish
        product = sale.product
        product.quantity += sale.quantity
        product.save()

        # Mijoz qarzini kamaytirish
        client = sale.client
        client.debt -= sale.debt_price
        client.save()

        sale.delete()
        return redirect('sales')


class ImportProductView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request):
        import_products = ImportProduct.objects.filter(branch=request.user.branch).order_by('-created_at', '-id')
        products = Product.objects.filter(branch=request.user.branch)

        # Search
        search = request.GET.get('search', '').strip()
        if search:
            import_products = import_products.filter(
                Q(product__name__icontains=search)
            )

        # Date filter
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        if date_from:
            import_products = import_products.filter(created_at__gte=date_from)
        if date_to:
            import_products = import_products.filter(created_at__lte=date_to)

        # Pagination
        paginator = Paginator(import_products, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'import_products': page_obj,
            'page_obj': page_obj,
            'products': products,
            'search': search,
            'date_from': date_from,
            'date_to': date_to,
        }

        return render(request, 'import_products.html', context=context)

    def post(self, request):
        if not request.user.branch:
            return render(request, 'warning.html', {
                'warning_title': 'Filial biriktirilmagan',
                'warning_message': "Sizning hisobingizga hech qanday filial (branch) biriktirilmagan. Iltimos, admin panel orqali filial biriktiring.",
                'back_url': 'imports'
            })

        product = get_object_or_404(Product, id=request.POST.get('product_id'))
        quantity = float(request.POST.get('quantity')) if request.POST.get('quantity') else 0.0
        buy_price = float(request.POST.get('buy_price')) if request.POST.get('buy_price') else 0.0
        
        ImportProduct.objects.create(
            product=product,
            buy_price=buy_price,
            quantity=quantity,
            user=request.user,
            branch=request.user.branch
        )
        product.quantity += quantity
        product.save()
        return redirect('imports')


class EditImportProductView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request, pk):
        import_product = get_object_or_404(ImportProduct, branch=request.user.branch, pk=pk)
        products = Product.objects.filter(branch=request.user.branch)

        context = {
            'import_product': import_product,
            'products': products,
        }
        return render(request, 'edit-imports.html', context=context)

    def post(self, request, pk):
        importproduct = get_object_or_404(ImportProduct, branch=request.user.branch, pk=pk)
        product = get_object_or_404(Product, branch=request.user.branch, id=request.POST.get("product_id"))

        def get_float(val):
            try: return float(val) if val else 0.0
            except: return 0.0

        new_quantity = get_float(request.POST.get('quantity'))
        buy_price = get_float(request.POST.get('buy_price'))

        # Eski miqdorni qaytarish
        old_product = importproduct.product
        old_quantity = importproduct.quantity
        old_product.quantity -= old_quantity
        old_product.save()

        importproduct.product = product
        importproduct.buy_price = buy_price
        importproduct.quantity = new_quantity
        importproduct.save()

        # Yangi miqdorni qo'shish
        product.quantity += new_quantity
        product.save()

        return redirect('imports')


class DeleteImportProductView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request, pk):
        import_product = get_object_or_404(ImportProduct, pk=pk, branch=request.user.branch)
        context = {
            'import_product': import_product,
        }
        return render(request, 'delete-imports.html', context=context)

    def post(self, request, pk):
        import_product = get_object_or_404(ImportProduct, pk=pk, branch=request.user.branch)
        import_product.delete()
        return redirect('imports')


class PaydebtView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request):
        pay_debts = PayDebt.objects.filter(branch=request.user.branch).order_by('-created_at', '-id')
        clients = Client.objects.filter(branch=request.user.branch).order_by('-created_at')

        # Search
        search = request.GET.get('search', '').strip()
        if search:
            pay_debts = pay_debts.filter(
                Q(client__name__icontains=search) | Q(client__shop_name__icontains=search)
            )

        # Date filter
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        if date_from:
            pay_debts = pay_debts.filter(created_at__gte=date_from)
        if date_to:
            pay_debts = pay_debts.filter(created_at__lte=date_to)

        # Pagination
        paginator = Paginator(pay_debts, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'pay_debts': page_obj,
            'page_obj': page_obj,
            'clients': clients,
            'search': search,
            'date_from': date_from,
            'date_to': date_to,
        }

        return render(request, 'pay-debts.html', context=context)

    def post(self, request):
        if not request.user.branch:
            return render(request, 'warning.html', {
                'warning_title': 'Filial biriktirilmagan',
                'warning_message': "Sizning hisobingizga hech qanday filial (branch) biriktirilmagan. Iltimos, admin panel orqali filial biriktiring.",
                'back_url': 'pay-debts'
            })

        client = get_object_or_404(Client, id=request.POST.get('client_id'))
        price = float(request.POST.get('price') or 0.0)

        if price == 0 or price > client.debt:
            return redirect('pay-debts')

        PayDebt.objects.create(
            client=client,
            price=price,
            description=request.POST.get('description'),
            branch=request.user.branch,
            user=request.user,
        )

        client.debt -= price
        client.save()
        return redirect('pay-debts')

class DownloadInvoiceView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request, pk):
        sale = get_object_or_404(Sale, pk=pk, branch=request.user.branch)
        pdf_buffer = generate_thermal_receipt(sale)
        
        response = FileResponse(
            pdf_buffer, 
            as_attachment=True, 
            filename=f"invoice_{sale.id}.pdf",
            content_type='application/pdf'
        )
        return response


class EditPayDebtView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request, pk):
        pay_debt = get_object_or_404(PayDebt, branch=request.user.branch, pk=pk)
        clients = Client.objects.filter(branch=request.user.branch)

        context = {
            'pay_debt': pay_debt,
            'clients': clients,
        }

        return render(request, 'edit-pay-debts.html', context=context)

    def post(self, request, pk):
        pay_debt = get_object_or_404(PayDebt, branch=request.user.branch, pk=pk)
        client = get_object_or_404(Client, branch=request.user.branch, id=request.POST.get('client_id'))

        pay_debt.price = request.POST.get('price')
        pay_debt.description = request.POST.get('description')
        pay_debt.client = client

        pay_debt.save()

        return redirect('pay-debts')


class DeletePayDebtView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request, pk):
        pay_debt = get_object_or_404(PayDebt, pk=pk, branch=request.user.branch)
        context = {
            'pay_debt': pay_debt,
        }
        return render(request, 'delete-pay-debt.html', context=context)

    def post(self, request, pk):
        pay_debt = get_object_or_404(PayDebt, pk=pk, branch=request.user.branch)
        pay_debt.delete()
        return redirect('pay-debts')


# ============================================
# EXCEL EXPORT VIEWS
# ============================================

class ExportProductsView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mahsulotlar"

        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['T/r', 'Nomi', 'Brendi', 'Narxi', 'Miqdori', "O'lchov", 'Jami qiymat']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        products = Product.objects.filter(branch=request.user.branch).order_by('name')
        for i, product in enumerate(products, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=product.name).border = thin_border
            ws.cell(row=row, column=3, value=product.brand or '-').border = thin_border
            ws.cell(row=row, column=4, value=product.price).border = thin_border
            ws.cell(row=row, column=5, value=product.quantity).border = thin_border
            ws.cell(row=row, column=6, value=product.unit).border = thin_border
            ws.cell(row=row, column=7, value=product.price * product.quantity).border = thin_border

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="mahsulotlar.xlsx"'
        wb.save(response)
        return response


class ExportSalesView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sotuvlar"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['T/r', 'Mahsulot', 'Mijoz', "Do'kon", 'Miqdori', 'Jami', "To'landi", 'Qoldi', 'Xodim', 'Sana']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        sales = Sale.objects.filter(branch=request.user.branch).order_by('-created_at')

        # Date filter
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        if date_from:
            sales = sales.filter(created_at__gte=date_from)
        if date_to:
            sales = sales.filter(created_at__lte=date_to)

        for i, sale in enumerate(sales, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=sale.product.name).border = thin_border
            ws.cell(row=row, column=3, value=sale.client.name).border = thin_border
            ws.cell(row=row, column=4, value=sale.client.shop_name).border = thin_border
            ws.cell(row=row, column=5, value=sale.quantity).border = thin_border
            ws.cell(row=row, column=6, value=sale.total_price).border = thin_border
            ws.cell(row=row, column=7, value=sale.paid_price).border = thin_border
            ws.cell(row=row, column=8, value=sale.debt_price).border = thin_border
            ws.cell(row=row, column=9, value=sale.user.first_name).border = thin_border
            ws.cell(row=row, column=10, value=str(sale.created_at)).border = thin_border

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sotuvlar.xlsx"'
        wb.save(response)
        return response


class ExportClientsView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mijozlar"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['T/r', 'FIO', "Do'kon nomi", 'Telefon', 'Manzili', 'Qarz']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        clients = Client.objects.filter(branch=request.user.branch).order_by('name')
        for i, client in enumerate(clients, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=client.name).border = thin_border
            ws.cell(row=row, column=3, value=client.shop_name).border = thin_border
            ws.cell(row=row, column=4, value=client.phone_number).border = thin_border
            ws.cell(row=row, column=5, value=client.address).border = thin_border
            ws.cell(row=row, column=6, value=client.debt).border = thin_border

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="mijozlar.xlsx"'
        wb.save(response)
        return response


class ExportImportsView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kirimlar"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['T/r', 'Mahsulot', 'Sotib olish narxi', 'Miqdori', 'Jami summa', 'Xodim', 'Sana']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        imports = ImportProduct.objects.filter(branch=request.user.branch).order_by('-created_at')

        # Date filter
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        if date_from:
            imports = imports.filter(created_at__gte=date_from)
        if date_to:
            imports = imports.filter(created_at__lte=date_to)

        for i, imp in enumerate(imports, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=imp.product.name).border = thin_border
            ws.cell(row=row, column=3, value=imp.buy_price).border = thin_border
            ws.cell(row=row, column=4, value=imp.quantity).border = thin_border
            ws.cell(row=row, column=5, value=imp.buy_price * imp.quantity).border = thin_border
            ws.cell(row=row, column=6, value=imp.user.first_name).border = thin_border
            ws.cell(row=row, column=7, value=str(imp.created_at)).border = thin_border

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="kirimlar.xlsx"'
        wb.save(response)
        return response


class ExportPayDebtsView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tolanilgan_qarzlar"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['T/r', 'Mijoz', 'Summa', 'Izoh', 'Xodim', 'Sana']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        pay_debts = PayDebt.objects.filter(branch=request.user.branch).order_by('-created_at')

        # Date filter
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        if date_from:
            pay_debts = pay_debts.filter(created_at__gte=date_from)
        if date_to:
            pay_debts = pay_debts.filter(created_at__lte=date_to)

        for i, pd in enumerate(pay_debts, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=pd.client.name).border = thin_border
            ws.cell(row=row, column=3, value=pd.price).border = thin_border
            ws.cell(row=row, column=4, value=pd.description or '-').border = thin_border
            ws.cell(row=row, column=5, value=pd.user.first_name).border = thin_border
            ws.cell(row=row, column=6, value=str(pd.created_at)).border = thin_border

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="tulangan_qarzlar.xlsx"'
        wb.save(response)
        return response


from .utils import generate_thermal_receipt
from django.http import FileResponse

class DownloadInvoiceView(LoginRequiredMixin, View):
    login_url = 'login'

    def get(self, request, pk):
        sale = get_object_or_404(Sale, pk=pk, branch=request.user.branch)
        pdf_buffer = generate_thermal_receipt(sale)
        
        response = FileResponse(
            pdf_buffer, 
            as_attachment=True, 
            filename=f"invoice_{sale.id}.pdf",
            content_type='application/pdf'
        )
        return response
